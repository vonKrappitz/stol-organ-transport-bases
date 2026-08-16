# Copyright 2026 Maciej M. Kasperek. Licensed under the Apache License, Version 2.0.
# paths relative to the repository root
import os as _os
try:
    _SCRIPT_DIR = _os.path.dirname(_os.path.abspath(__file__))
except NameError:
    _SCRIPT_DIR = _os.path.dirname(_os.path.abspath(_os.sys.argv[0])) or _os.getcwd()
_REPO_DIR  = _os.path.dirname(_SCRIPT_DIR)          # repository root
_DATA_DIR   = _os.path.join(_REPO_DIR, "data")
RESULTS      = _os.path.join(_REPO_DIR, "results")     # figures and CSV go here
_os.makedirs(RESULTS, exist_ok=True)
_os.chdir(_DATA_DIR)                                  # bare paths read from data/
# end of path block

# paths; NFZ data handled gracefully when absent
import os as _os2, sys as _sys2
_nfz = _os2.path.join(_DATA_DIR, "nhf_interhospital_transport_2020_2024.xlsx")
_nfzc = _os2.path.join(_DATA_DIR, "nhf_transport_od.csv")
if not _os2.path.exists(_nfz) and not _os2.path.exists(_nfzc):
    print("=" * 68)
    print("National Health Fund data file not found in data/.")
    print("See data/NHF_transport_data.md for the file name and provenance")
    print("The core results reproduce without them from the other scripts. Skipping.")
    print("=" * 68)
    _sys2.exit(0)
# --- end obslugi NFZ ---

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# 16_transport_flows.py
# Wejscie: data/nfz_transport_od.xlsx (kopia file NFZ).
# step 1. Konwersja do tidy CSV data/nhf_transport_od.csv.
# Step 2. Inter-voivodeship flow matrix by decision:
#   D1 cenzura "<5" -> 2 (variants 1 i 4 do wrazliwosci),
#   D2 oba produkty total (naziemny + lotniczy),
#   D3 sum 2020-2024 (variant 2023-2024 do wrazliwosci).
# Step 3. Flow distribution by centroid distance (haversine as in 04_minimax).
import openpyxl, json, math, csv, sys
from collections import defaultdict

XLSX = "nhf_interhospital_transport_2020_2024.xlsx"
CSV  = "nhf_transport_od.csv"

def hav(lat1,lon1,lat2,lon2):
    R=6371.0; p1,p2=math.radians(lat1),math.radians(lat2)
    dp=math.radians(lat2-lat1); dl=math.radians(lon2-lon1)
    a=math.sin(dp/2)**2+math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return 2*R*math.asin(math.sqrt(a))

# --- step 1, xlsx -> tidy csv ---
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb['Tabela_1']
rows=[]
for r in range(3, ws.max_row+1):
    v=[ws.cell(row=r,column=c).value for c in range(1,7)]
    if v[0] is None: continue
    rok, wy, pr, kod, naz, n = v
    cenz = isinstance(n,str)
    rows.append((int(rok), wy.strip(), pr.strip(), kod.strip(), int(n) if not cenz else '', int(cenz)))
assert len(rows)==680, f"wierszy {len(rows)}, oczekiwano 680"
with open(CSV,'w',newline='',encoding='utf-8') as f:
    w=csv.writer(f)
    w.writerow(['year','voivodeship_from','voivodeship_to','product_code','services','censored_below_5'])
    w.writerows(rows)
print(f"CSV saved: {CSV}, rows {len(rows)}")
n_cenz=sum(r[5] for r in rows)
print(f"censored: {n_cenz}, numeric: {len(rows)-n_cenz}")

# --- step 2, matrix under rules D1-D3 ---
CENZ = 2  # D1
def macierz(lata, cenz_val=CENZ):
    m=defaultdict(int)
    for rok,wy,pr,kod,n,c in rows:
        if rok not in lata: continue
        if wy==pr: continue  # STOL nie dotyczy ruchu wewnatrz voivodeships
        m[(wy,pr)] += cenz_val if c else n
    return m

M_full  = macierz({2020,2021,2022,2023,2024})
M_post  = macierz({2023,2024})
M_lo    = macierz({2020,2021,2022,2023,2024}, cenz_val=1)
M_hi    = macierz({2020,2021,2022,2023,2024}, cenz_val=4)
print(f"\npairs in main matrix: {len(M_full)}")
print(f"main flow total (cens=2): {sum(M_full.values())}")
print(f"censoring brackets: [{sum(M_lo.values())}, {sum(M_hi.values())}]")
print(f"total 2023-2024 (cens=2): {sum(M_post.values())}")

# --- step 3, distances i distribution ---
geo=json.load(open('voivodeship_boundaries.json', encoding="utf-8"))
cent=geo['cent']
def klucz(w):  # nazwy NFZ -> klucze voivodeship_boundaries.json
    return w.lower()
brak=[w for w in set(k for p in M_full for k in p) if klucz(w) not in cent]
assert not brak, f"brak centroidow dla: {brak}"

pasma=[(0,100),(100,200),(200,300),(300,400),(400,1000)]
suma=sum(M_full.values())
print("\ninter-voivodeship flows by centroid distance (main matrix):")
print(f"{'band km':>12} {'flows':>10} {'share':>8} {'cumul.':>8} {'pairs':>5}")
skum=0
for lo,hi in pasma:
    s=0; npar=0
    for (wy,pr),v in M_full.items():
        d=hav(*cent[klucz(wy)], *cent[klucz(pr)])
        if lo<=d<hi: s+=v; npar+=1
    skum+=s
    print(f"{lo:>5}-{hi:<6} {s:>10} {100*s/suma:>7.1f}% {100*skum/suma:>7.1f}% {npar:>5}")

print("\npairs above 300 km centroid to centroid, flows (cenz=2, 5 lat):")
dl=[]
for (wy,pr),v in M_full.items():
    d=hav(*cent[klucz(wy)], *cent[klucz(pr)])
    if d>=300: dl.append((d,wy,pr,v))
for d,wy,pr,v in sorted(dl, key=lambda t:-t[3])[:15]:
    print(f"  {wy:>22} -> {pr:<22} {d:5.0f} km {v:>5}")
print(f"pairs >=300 km with flow: {len(dl)}, flow total: {sum(t[3] for t in dl)}")

# flow-weighted median distance
wd=[]
for (wy,pr),v in M_full.items():
    d=hav(*cent[klucz(wy)], *cent[klucz(pr)])
    wd += [d]*v
wd.sort()
print(f"\nflow-weighted median distance: {wd[len(wd)//2]:.0f} km")
print(f"weighted mean: {sum(wd)/len(wd):.0f} km")
